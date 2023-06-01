import openpyxl
from utils import PublicKeyUtils
from aptos_sdk.account import Account
from mnemonic import Mnemonic

# Создаем новую книгу и выбираем активный лист
workbook = openpyxl.Workbook()
sheet = workbook.active

num_wallets = 100  # Укажите количество кошельков, которые вы хотите сгенерировать

for i in range(num_wallets):
    seed_phrase = Mnemonic('english').generate()
    pt_seed = PublicKeyUtils(seed_phrase)
    keys = Account.load_key(pt_seed.private_key.hex())
    address = keys.address()
    public_key = keys.public_key()
    private_key = keys.private_key.hex()

    print(f"Mnemonic: {seed_phrase}")
    print(f"Address: {address}")
    print(f"Public Key: {public_key}")
    print(f"Private Key: 0x{private_key}\n")

    # Записываем данные в ячейки Excel
    sheet.cell(row=i+1, column=1, value=seed_phrase)
    sheet.cell(row=i+1, column=2, value=str(address))
    sheet.cell(row=i+1, column=3, value=str(public_key))
    sheet.cell(row=i+1, column=4, value=private_key)

# Сохраняем файл после записи всех данных
workbook.save('martianAptosNEW.xlsx')
