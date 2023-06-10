from generate_wallet import generate_wallets
from aptos_data import get_aptos_wallet_info
from sui_data import get_sui_wallet_info
import openpyxl

num_wallets = int(input("Введите количество кошельков для генерации: "))

wallet_phrases = generate_wallets(num_wallets)

aptos_wallet_info = get_aptos_wallet_info(wallet_phrases)
sui_wallet_info = get_sui_wallet_info(wallet_phrases)

workbook = openpyxl.Workbook()
sheet = workbook.active

headers = ["Seed Phrase", "Aptos Address", "Aptos Private Key", "Sui Address", "Sui Private Key"]
sheet.append(headers)

for i, info in enumerate(aptos_wallet_info, start=2):
    sheet.cell(row=i, column=1, value=info['Seed Phrase'])
    sheet.cell(row=i, column=2, value=str(info['Address']))
    sheet.cell(row=i, column=3, value=str(info['Private Key']))

for i, info in enumerate(sui_wallet_info, start=2):
    sheet.cell(row=i, column=4, value=str(info['Address']))
    sheet.cell(row=i, column=5, value=str(info['Private Key']))

workbook.save('wallets.xlsx')

with open('seeds.txt', 'w') as f:
    for wallet in wallet_phrases:
        f.write(f"{wallet}\n")

with open('aptos_addresses.txt', 'w') as f:
    for info in aptos_wallet_info:
        f.write(f"{info['Address']}\n")

with open('aptos_private_keys.txt', 'w') as f:
    for info in aptos_wallet_info:
        f.write(f"{info['Private Key']}\n")

with open('sui_addresses.txt', 'w') as f:
    for info in sui_wallet_info:
        f.write(f"{info['Address']}\n")

with open('sui_private_keys.txt', 'w') as f:
    for info in sui_wallet_info:
        f.write(f"{info['Private Key']}\n")

print(f"{num_wallets} Сгенерированы. Результаты сохранены в файлах: seeds.txt, aptos_addresses.txt, aptos_private_keys.txt, sui_addresses.txt, sui_private_keys.txt и wallets.xlsx.")
